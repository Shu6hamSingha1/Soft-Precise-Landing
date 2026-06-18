#!/usr/bin/env python3
"""Build MATLAB_test_record.xlsx — a test-records spreadsheet for the MATLAB
c-tilde-h (option-b) investigation, mirroring PX4_Gazebo/test_data/Landing_Test/
parameter_record.ods. Re-run to regenerate.  python3 build_test_record.py
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
hdrfill = PatternFill("solid", fgColor="1F4E78"); hdrfont = Font(bold=True, color="FFFFFF", size=10)
vpass = Font(size=9, color="006100"); vfail = Font(size=9, color="9C0006")
thin = Side(style="thin", color="BBBBBB"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")

def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c); cell.fill = hdrfill; cell.font = hdrfont
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 28

# ===== Sheet 1: c-tilde-h test log =====
ws = wb.active; ws.title = "cthilde_test_log"
cols = ["Trial", "Date", "Harness (.m)", "What varied", "Config (fixed)", "IC / Traj",
        "n", "Hypothesis", "Result", "Verdict"]
ws.append(cols)
rows = [
 ["MT1", "2026-06-15", "csimple_ic5", "c-term: old full-w vs C_SIMPLE (simple c-tilde-h)", "K_rd=1.4375, GAMMA_COG=0.005", "IC5 Circular", "12 noisy", "corrected c-tilde-h is cleaner/less noisy -> should not regress", "old-c SP 10/12; C_SIMPLE SP 6/12 (fail 7,8 marginal + seed8 divergence)", "REGRESSES"],
 ["MT2", "2026-06-15", "(seed-8 capture)", "diagnose C_SIMPLE seed-8 fail", "K_rd=1.4375", "IC5 Circular seed8", "1", "find why C_SIMPLE diverges where old-c lands", "identical to t4.6; then kappa_z runaway 0.19->2.99 at touchdown; kappa_z clamp 0.5 did NOT fix", "Terminal/CBF wall tip-over"],
 ["MT3", "2026-06-15", "csimple_retune", "K_rd in {1.4375,2.5,4.0,6.0}", "C_SIMPLE=1", "IC5 Circular sd 1,5,10,11,12", "5 noisy", "terminal lateral damping recovers the marginal-softness seeds", "K_rd=2.5 -> 5/5 SP; 4.0 over-damps (1/5); 6.0 diverges", "K_rd=2.5 sweet spot (narrow)"],
 ["MT4", "2026-06-15", "gate_csimple", "old-c base vs C_SIMPLE+K_rd2.5", "GAMMA_COG=0.005", "IC1-5 Circular", "nl1 + ny5", "corrected controller passes IC1-5 no-regression gate", "old-c 29/30; C_SIMPLE+K_rd2.5 30/30; NO regression; IC5 noisy 4/5->5/5", "PASSES gate (Circular)"],
 ["MT5", "2026-06-15", "csimple_final", "old-c vs C_SIMPLE+K_rd2.5 (full)", "--", "IC5 Circular", "12 noisy", "confirm broader IC5 picture incl. seed8", "old-c 10/12; C_SIMPLE+K_rd2.5 9/12 -- ELIMINATES fly-aways (seed4 v1.43->0.094, seed8 v2.07->0.26), softer mean", "~equivalent, safer; -1 at cutoff"],
 ["MT6", "2026-06-16", "csimple_krd_refine", "K_rd in {2.8,3.0}", "C_SIMPLE=1", "IC5 Circular", "12 noisy", "push K_rd to recover the 2 marginal misses", "ALL K_rd {2.5,2.8,3.0} = 9/12; misses SHUFFLE seeds; 3.0 over-damps", "9/12 ceiling; keep K_rd=2.5"],
 ["MT7", "2026-06-16", "test_new_formulation", "old-c vs C_SIMPLE+K_rd2.5", "TRAJ_OVERRIDE hook", "5 traj x 5 IC", "1 noisy", "new formulation holds across all trajectories", "old-c 24/25; C_SIMPLE+K_rd2.5 23/25; equal on Lin/Sin/Liss/Circ; regress Static IC5", "EQUIVALENT 4/5 traj; -1 at IC5"],
 ["MT8", "2026-06-16", "test_static_ic5", "old-c vs C_SIMPLE (K_rd 1.44 & 2.5)", "--", "Static IC5", "5 noisy", "is the Static IC5 regression robust + is K_rd the cause", "old-c 3/5; C_SIMPLE+K_rd2.5 2/5; K_rd1.44 1/5 -- both fail sd 3,4 (shared wall)", "Real ~1-seed deficit; K_rd NOT cause"],
 ["MT9", "2026-06-16", "recover_ic5", "reaching gain Gamma_xy in {0.5,1.5,3.0}", "C_SIMPLE+K_rd2.5", "Static+Circular IC5", "5 noisy", "more SMC aggressiveness recovers the lateral-closure stall", "Gamma_xy up WORSE on both (Static 2->1->0; Circ 5->1->1)", "FAIL - destabilizes"],
 ["MT10", "2026-06-16", "sen_recover_sweep", "SEN_RECOVER_ST in {0,0.3,0.5,0.7}", "C_SIMPLE+K_rd2.5", "Static+Circular IC5", "5 noisy", "S9 hard-containment forces s_e_n inside p_s -> recover", "ZERO effect (Static 2/5 at all ST) -- triggers on raw_ratio>=1 but stall is at the S_s=0.95 clamp (pre-breach)", "FAIL - fires too late; inner-bound"],
 ["MT11", "2026-06-16", "sen_psw_sweep", "SEN funnel width p_s_0 in {1.2,2.0,3.0}  [NOTE: p_s_0 is FoV-LOCKED, not a usable knob]", "C_SIMPLE+K_rd2.5", "Static+Circular IC5", "5 noisy", "wider funnel avoids the demand-starvation clamp", "WORSE on both (Static 2->1; Circ 5->3->2) -- looser funnel = weaker barrier", "FAIL + lever is locked (data only)"],
 ["MT12", "2026-06-16", "sen_krp_sweep", "SEN proportional K_rp in {9,11,13}", "C_SIMPLE+K_rd2.5", "Static+Circular IC5", "5 noisy", "stronger closing demand -> faster early closure", "no help Static (2/5 at 11, 1/5 at 13); hurts Circular (5->4)", "FAIL - G_s^-1 collapse kills it at clamp"],
 ["MT13", "2026-06-16", "sen_mid_sweep", "middle-loop E_xy 0.5 / kappa0_xy 0.5", "C_SIMPLE+K_rd2.5", "Static+Circular IC5", "5 noisy", "stiffer/stronger middle loop delivers the closing motion", "no help Static (2/5); E0.5+k0.5 worse on both", "FAIL - intrinsic, not a delivery gain"],
 ["MT14", "2026-06-16", "sen_p20_sweep", "velocity-funnel width p_20_xy in {10,25,40}  (p_20 unlocked)", "C_SIMPLE+K_rd2.5", "Static+Circular IC5", "5 noisy", "tighter velocity funnel -> middle-loop barrier bites earlier -> delivers closing motion", "p20=10 RECOVERS Static IC5 2/5->4/5 (exceeds old-c 3/5), Circular 5/5; p20=40 worse", "RECOVERY FOUND - tighter p_20"],
 ["MT15", "2026-06-16", "test_p20_suite", "full-suite validation of p20=10", "C_SIMPLE+K_rd2.5+p20=10", "5 traj x 5 IC", "1 noisy", "p20=10 recovers IC5 WITHOUT breaking other traj/IC", "PENDING (running)", "PENDING"],
]
for r in rows:
    ws.append(r)
for i, w in enumerate([7, 11, 17, 34, 22, 18, 8, 34, 50, 28], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
style_header(ws, len(cols))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(cols) + 1):
        cell = ws.cell(r, c); cell.alignment = wrap; cell.border = border; cell.font = Font(size=9)
    v = (ws.cell(r, 10).value or "")
    if "PASS" in v or "equivalent" in v or "sweet spot" in v or "safer" in v:
        ws.cell(r, 10).font = vpass
    elif "FAIL" in v or "REGRESS" in v:
        ws.cell(r, 10).font = vfail

# ===== Sheet 2: findings & conclusions =====
ws2 = wb.create_sheet("findings")
ws2.append(["#", "Finding", "Evidence", "Implication"])
f = [
 ["F1", "Corrected manuscript c-term (C_SIMPLE) is the analytically-correct kinematics (transport theorem) and less noisy (clean IMU psi-dot-b vs noisy optic-flow w/w-dot).", "manuscript.tex L196/L208/L212; old-form residual 2.5-4x GT vs ~2% corrected", "The MODEL is right; only the closed-loop behaviour is in question"],
 ["F2", "C_SIMPLE passes the Circular IC1-5 gate (30/30) but is ~1 seed worse on IC5 [2,2,-3] across ALL trajectories.", "MT4 (30/30) vs MT7 (Static IC5 regress); MT8 (3/5->2/5)", "Option (b) no-regression premise holds only on Circular, fails at IC5"],
 ["F3", "s_e_n BOUNDEDNESS: under C_SIMPLE, s_e_n_y pins at the S_s=-0.95 clamp then BREACHES p_s (->6x, 46% of steps, no convergence); old-c stays bounded (max S_s 0.88) and converges.", "Static IC5 seed1 capture (V_s_e_n vs p_s, residency S_s)", "The corrected formulation LACKS the control authority the funnel guarantee needs"],
 ["F4", "Mechanism: old-c's INCORRECT full-w terms create a productive sustained flow error (|V_h_e|~1.0-1.7) that keeps driving closure; the leaner-correct c-term tracks cleanly (|V_h_e|~0.5-0.9) and under-closes.", "Static IC5 old-c vs C_SIMPLE trajectory compare", "The old robustness was partly an artifact of the wrong feedforward -- not replicable principled-ly"],
 ["F5", "IC5 deficit IS inner-bound (middle-loop delivery): outer/SEN levers (K_rp, p_s_0, SEN_RECOVER_ST) + reaching/middle gains (Gamma_xy, E_xy, kappa0) all fail. The fix is TIGHTER p_20 (velocity-funnel width 25->10): barrier bites earlier -> more aggressive flow tracking -> delivers the closing motion.", "MT9-MT13 fail; MT14 p20=10 recovers Static IC5 2/5->4/5", "Option (b) RECOVERABLE via p_20 (pending full-suite validation MT15)"],
 ["F6", "RECOMMENDATION (UPDATED): if MT15 confirms p20=10 holds across the full suite without regression, OPTION (b) is back on (corrected model + matching controller C_SIMPLE+K_rd2.5+p20=10 + regenerated results). Else fall back to option (c) (corrected model, keep old-code results, residual->d_h).", "MT14 breakthrough", "Validate p20=10 full suite + IC1-5 gate before baking"],
 ["F7", "CONSTRAINT (corrected 2026-06-16): p_s_0 (SEN/position funnel initial width) is LOCKED -- FoV-dependent (the funnel admits the normalized position error s_e_n = s_e/p_10, bounded by the sensor/FoV half). p_20 (velocity funnel) IS free to retune.", "user 2026-06-16", "Exclude p_s_0 from any retune; p_20 is fair game"],
 ["F8", "SUPERSEDED BY COMBINED-BARRIER (2026-06-18): rather than tune the back-mapped form, IMPLEMENTED the combined/blended sliding surface of control_formulation.tex (position barrier zeta_r DIRECTLY in sigma: sigma_xy=zeta_h+chi_r*zeta_r; h_d=measured-s_dot+transport+h_rd*s, no back-map). No G_s^-1 demand-starvation. Gated COMBINED_BARRIER (default-off).", "CB1-CB17 (combined_barrier_log sheet)", "Structural fix > tuning; option (b) unblocked via direct position authority"],
 ["F9", "Two DERIVATION-DRIVEN fixes made it work: (a) s_ddot-drop -- c_h=c_tilde_h-h_d_dot, and h_d_dot of the measured-s_dot h_d carries the centroid accel s_ddot, which is 1/z-inflated -> over-aggressive terminal -> drop it into d_h for kappa. (b) p_2inf_xy -- the flow-funnel terminal floor bounds the terminal flow error = chase-lag velocity.", "CB4 (s_ddot derivation+fix), CB7 (p_2inf lever)", "s_ddot belongs mathematically but is dynamically over-aggressive; p_2inf is the principled velocity lever"],
 ["F10", "chi_r (surface gain in sigma=zeta_h+chi_r*zeta_r) is the position-vs-velocity weight. On sigma=0: zeta_h=-chi_r*zeta_r, so position offset ~ 1/chi_r (INTERIOR, not funnel-limited -> p_r_inf can't fix it). Lower chi_r = velocity damping (chase-lag) at the cost of precision. chi_r=0.65 threads both opposed margin cells.", "CB10-CB14; Circ IC4 offset interior resid 0.26", "chi_r is the binding trade knob; 0.65 = the thread point"],
 ["F11", "WINNING CONFIG: COMBINED_BARRIER + corrected c_tilde_h + s_ddot-drop + p_2inf_xy=0.5 + chi_r=0.65 -> 25/25 SP noiseless, 75/75 SP noisy (3 seeds), worst funnel residency 0.665 (NO breach). BEATS the back-mapped form (24/25).", "CB14 (25/25), CB16 (75/75 noisy)", "Combined-barrier is the better controller AND aligns code<->paper"],
 ["F12", "PROOF CONSTRAINT (COMBINED_SURFACE_PROOF_ADDENDUM, Ubuntu pull 74ed0ee): the funnel guarantee needs p_r_inf >= 1 (FoV-consistent, Standing Condition 1) -- precision via (p_h, chi_r), NOT sub-FoV p_r. Implementation used p_r_inf=0.40 -> reconcile to 1.0.", "CB17 (p_r_inf=1.0 reconcile)", "Align p_r_inf>=1 with the proof; precision unaffected (comes from p_h/chi_r)"],
]
for r in f:
    ws2.append(r)
for i, w in enumerate([6, 54, 46, 46], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
style_header(ws2, 4)
for r in range(2, ws2.max_row + 1):
    for c in range(1, 5):
        cell = ws2.cell(r, c); cell.alignment = wrap; cell.border = border; cell.font = Font(size=9)

# ===== Sheet 3: combined-barrier trial log =====
ws3 = wb.create_sheet("combined_barrier_log")
cb_cols = ["Trial", "Date", "Harness (.m)", "What varied", "Result", "Verdict"]
ws3.append(cb_cols)
cb = [
 ["CB1", "2026-06-18", "(port + sanity)", "gated COMBINED_BARRIER on; both c_tilde_h forms, default params", "BOTH diverge early (FoV fail) -- h_d had V_sd=0 (no s_dot feedforward)", "h_d bug found"],
 ["CB2", "2026-06-18", "(h_d fix)", "h_d = MEASURED finite-diff s_dot + transport + h_rd*s", "corrected c_tilde_h LANDS Circ IC2 (xy0.046 precise); old-c_tilde_h diverges y -> corrected is the pairing", "no more divergence at IC2"],
 ["CB3", "2026-06-18", "cb_chi/ph_sweep", "chi_r {0.05-1}, p_h {1.5-8} sweeps", "ALL diverge (undamped overshoot) -- revealed the s_ddot issue in h_d_dot", "single-funnel tuning can't fix it"],
 ["CB4", "2026-06-18", "cb_sddot_test", "DERIVATION: c_h=c_tilde_h-h_d_dot carries centroid accel s_ddot (1/z-inflated). CB_DROP_SDDOT -> d_h", "IC5 Static/Circ diverge -> LAND (xy0.057/0.012); r_bar_e bounded (resid0.59-0.67)", "s_ddot-drop FIXES terminal divergence"],
 ["CB5", "2026-06-18", "cb_suite", "s_ddot-drop full 5x5 (corrected, chi_r=1)", "25/25 land, 21/25 SP; 4 Pr = chase-lag terminal velocity", "combined-barrier WORKS; chase-lag residual"],
 ["CB6", "2026-06-18", "cb_filt+suite_tau", "s_ddot FILTER (CB_SDDOT_TAU=0.5) vs drop", "20/25 (net WASH): helps Sin IC3/Circ IC5 but regresses the IC3s; LPF phase-lag", "filter = blunt global knob, worse than drop"],
 ["CB7", "2026-06-18", "cb_p2inf_sweep+suite", "p_2inf_xy (flow-funnel TERMINAL floor) 2.5->0.5", "24/25; clean MONOTONIC vel drop, controls IMPROVE; the principled chase-lag lever", "p_2inf_xy = the velocity lever"],
 ["CB8", "2026-06-18", "cb_breach_check", "user directive: ensure r_bar_e (s_e_n) doesn't breach p_r (p_s)", "worst residency 0.674 < 1 across ALL 25 -- NO breach; prescribed-performance bound holds", "guarantee intact"],
 ["CB9", "2026-06-18", "cb_yaxis", "per-axis vel; p_2inf_y only", "Liss IC3 vy floors ~0.19 (can't cross 0.20 via p_2inf); vy-dominant chase", "p_2inf alone leaves Liss IC3 at boundary"],
 ["CB10", "2026-06-18", "cb_other_params", "Gamma/gamma2/chi_r/E on Liss IC3", "chi_r=0.5 -> vel0.103, gamma2=0.5 -> 0.154 close it; DIRECTION: lower chi_r / faster gamma2 help; higher DIVERGE", "chi_r & gamma2 are bandwidth levers"],
 ["CB11", "2026-06-18", "cb_suite (chi_r=0.5)", "chi_r=0.5 full suite", "24/25 -- closes Liss IC3 but Circ IC4 loses PRECISION (lower chi_r de-weights zeta_r)", "chi_r trades velocity for precision"],
 ["CB12", "2026-06-18", "cb_suite_g2", "gamma2=0.5 full suite", "20/25 -- wrecks the Circular trajectory (over-fast contraction)", "gamma2 out"],
 ["CB13", "2026-06-18", "cb_combo", "chi_r x p_r_inf combo", "p_r_inf CAN'T bridge: Circ IC4 offset is INTERIOR (resid 0.26, chi_r-set), not funnel-limited", "confirms WHY chi_r=0.5 misses precision"],
 ["CB14", "2026-06-18", "cb_chi_fine+suite_065", "chi_r=0.65 (the thread point)", "25/25 SP, NO breach (0.658); both opposed margin cells just under 0.20 (0.196/0.192)", "CLEAN 25/25 -- beats back-mapped 24/25"],
 ["CB15", "2026-06-18", "cb_vpcomp", "component breakdown of margin cells", "Liss IC3 binds vy(-0.173), Circ IC4 binds vx(+0.149) -- DIFFERENT axes; asymmetric-ready", "per-axis structure mapped"],
 ["CB16", "2026-06-18", "cb_noisy", "NOISY validation (chi_r0.65 config), 3 seeds x 5x5", "75/75 SP, 75/75 land, worst residency 0.665 -- ROBUST under pixel noise", "noise-robust, not noiseless-lucky"],
 ["CB17", "2026-06-18", "cb_prinf1", "PROOF reconcile: p_r_inf=1.0 (Standing Condition 1, FoV-consistent)", "25/25 SP, residency 0.615 (tighter); precision via (p_h,chi_r) per proof", "code aligned with the Lyapunov proof"],
 ["CB18", "2026-06-18", "cb_proof_guided+suite", "PROOF reconcile p_r_inf=1.0 (Standing Cond 1) + manifold-guided chi_r", "p_r_inf=1.0 -> still 25/25, WIDENED margins (Liss IC3 0.196->0.136); chi_r=0.85 BEATS 0.65 (binding margin 0.180 vs 0.192); chi_r=1.0 breaches (resid 1.282, 24/25)", "PROOF improved SP margin; chi_r=0.85 optimal"],
 ["CB19", "2026-06-18", "(bake + cb_12seed)", "FINAL: bake combined defaults (chi_r=0.85, p_r_inf=1.0, p_2inf=0.5, s_ddot-drop default-on)", "baked so COMBINED_BARRIER=1 gives 25/25 OOTB", "FINAL CONFIG BAKED"],
 ["CB20", "2026-06-18", "cb_12seed", "12-seed noisy (5x5x12=300) on the drop config", "205/300 SP (68%), 268/300 land, worstResid 5.3 -- looks noise-fragile; the 3-seed 75/75 was optimistic", "noise picture needs context (CB21)"],
 ["CB21", "2026-06-18", "cb_vs_bm", "NOISE: combined-barrier vs BACK-MAPPED, 4 hard cells x 6 seeds", "combined 16/24 BEATS back-mapped 13/24; far less breach (2.83 vs 19.29) -- the SEN G_s^-1 starvation is worse under noise", "NOT a noise regression -- combined-barrier comparably-to-MORE robust"],
 ["CB22", "2026-06-18", "cb_sddot_kept/tuned", "KEEP s_ddot: hard-cap (DH_D_CAP), tau-LPF (CB_SDDOT_TAU), SG-clean (CB_SDOT_FILT)", "ALL worse than drop: cap2 still un-lands IC5; SG-clean DIVERGES (full s_ddot magnitude over-drives); tau0.8 best at 5/6", "keep-s_ddot strictly worse than drop"],
 ["CB23", "2026-06-18", "cb_filt_noisy", "Savitzky-Golay s_dot filter (causal quadratic least-squares) under noise", "16/24 UNCHANGED across W=0/9/13 (worstResid 2.83->2.58 only) -- s_dot is NOT the dominant noise path; the measured flow h is", "SG filter kept as gated option; not the noise fix"],
 ["CB24", "2026-06-18", "cb_peraxis/asym/rectify/other_gains", "s_ddot-KEPT per-axis tuning: chi_r=[1.1,0.65] + rectifiers", "22/25, breaches: Circ IC4 NEEDS chi_r_x>=1.1 (precision) but chi_r_x>=1.1 destabilizes Static/Liss IC3 (Y-breach, X-tilt cross-couples to Y); NO p_2inf/E/chi_r_y rectifies", "IRRECONCILABLE chi_r_x conflict; root = the 1/z-inflated s_ddot"],
]
for r in cb:
    ws3.append(r)
for i, w in enumerate([7, 11, 20, 40, 50, 30], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
style_header(ws3, len(cb_cols))
for r in range(2, ws3.max_row + 1):
    for c in range(1, len(cb_cols) + 1):
        cell = ws3.cell(r, c); cell.alignment = wrap; cell.border = border; cell.font = Font(size=9)
    v = (ws3.cell(r, 6).value or "")
    if "FIX" in v or "WORKS" in v or "25/25" in v or "lever" in v or "robust" in v or "aligned" in v or "intact" in v:
        ws3.cell(r, 6).font = vpass
    elif "out" in v or "wash" in v or "bug" in v or "trades" in v:
        ws3.cell(r, 6).font = vfail

import os
out = os.path.join(os.path.dirname(__file__), "MATLAB_test_record.xlsx")
wb.save(out)
print("wrote", out, "-", ws.max_row - 1, "trials,", ws2.max_row - 1, "findings")
